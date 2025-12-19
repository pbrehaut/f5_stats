import json
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Mapping of host IDs to sites
CHASSIS_MAPPING = {
    'chs412276s': 'HMB-Viprion-01',
    'chs412821s': 'HMB-Viprion-02',
    'chs412274s': 'GSW-Viprion-01',
    'chs412822s': 'GSW-Viprion-02',
    'f5-arut-orvq': 'HMB-F5-r5900-03',
    'f5-buoc-hruj': 'HMB-F5-r5900-04',
}


def extract_file_info(filepath):
    """Extract host ID and timestamp from filename"""
    filename = os.path.basename(filepath)
    filename_without_ext = filename.replace('.json', '')

    # Expected format: hostID__hostname__YYYY-MM-DD__HH-MM-SS__type_pool_members
    parts = filename_without_ext.split('__')

    site_id = parts[0]
    host_name = parts[1]
    date_str = parts[2]
    time_str = parts[3]

    # Map host ID to alternate hostname
    alt_site_name = CHASSIS_MAPPING.get(site_id, site_id)

    # Parse datetime using datetime module
    datetime_str = "{} {}".format(date_str, time_str.replace('-', ':'))
    return alt_site_name + ' ' + host_name, datetime_str


def get_available_files(pool_type):
    """Scan current directory for pool_members.json files"""
    suffix = '__{}_pool_members.json'.format(pool_type)
    files = [f for f in os.listdir('.') if f.endswith(suffix)]
    return sorted(files)


def display_menu(files):
    """Display menu with mapped site names"""
    print("\nAvailable files:")
    print("-" * 80)
    for idx, filepath in enumerate(files, 1):
        site_name, timestamp = extract_file_info(filepath)
        print("{}. {} - {}".format(idx, site_name, timestamp))
    print("-" * 80)


def select_pool_type():
    """Allow user to select pool type"""
    print("\nSelect pool type:")
    print("-" * 80)
    print("1. LTM (Local Traffic Manager)")
    print("2. GTM (Global Traffic Manager)")
    print("-" * 80)

    while True:
        try:
            choice = input("\nSelect pool type (enter number): ").strip()
            if choice == '1':
                return 'ltm'
            elif choice == '2':
                return 'gtm'
            else:
                print("Invalid selection. Please enter 1 or 2.")
        except KeyboardInterrupt:
            print("\nOperation cancelled.")
            return None


def select_files(pool_type):
    """Allow user to select exactly 2 files"""
    files = get_available_files(pool_type)

    if len(files) < 2:
        print("Error: At least 2 {} files are required for comparison.".format(pool_type.upper()))
        return None, None

    display_menu(files)

    selected_files = []

    while len(selected_files) < 2:
        try:
            choice = input("\nSelect file {} (enter number): ".format(len(selected_files) + 1)).strip()
            idx = int(choice) - 1

            if idx < 0 or idx >= len(files):
                print("Invalid selection. Please try again.")
                continue

            if files[idx] in selected_files:
                print("File already selected. Please choose a different file.")
                continue

            selected_files.append(files[idx])

        except ValueError:
            print("Please enter a valid number.")
        except KeyboardInterrupt:
            print("\nOperation cancelled.")
            return None, None

    return selected_files[0], selected_files[1]


def compare_pool_states(file1_path, file2_path, output_path):
    # Extract file information
    host1, time1 = extract_file_info(file1_path)
    host2, time2 = extract_file_info(file2_path)

    # Load JSON files
    with open(file1_path, 'r') as f:
        data1 = json.load(f)

    with open(file2_path, 'r') as f:
        data2 = json.load(f)

    results = []
    mismatches = []

    # Column headers with host and time information
    col1_header = "{} ({})".format(host1, time1)
    col2_header = "{} ({})".format(host2, time2)

    # Get all pool names from both files
    all_pools = set(data1.keys()) | set(data2.keys())

    for pool_name in sorted(all_pools):
        # Compare pool level
        pool1_state = data1.get(pool_name, {}).get('status.availability-state', 'N/A')
        pool2_state = data2.get(pool_name, {}).get('status.availability-state', 'N/A')

        match = 'Yes' if pool1_state == pool2_state else 'No'

        results.append({
            'Type': 'Pool',
            'Name': pool_name,
            col1_header: pool1_state,
            col2_header: pool2_state,
            'Match': match
        })

        if match == 'No':
            mismatches.append({
                'Type': 'Pool',
                'Name': pool_name,
                col1_header: pool1_state,
                col2_header: pool2_state
            })

        # Get all members from both files for this pool
        members1 = data1.get(pool_name, {}).get('members', {})
        members2 = data2.get(pool_name, {}).get('members', {})
        all_members = set(members1.keys()) | set(members2.keys())

        # Compare members
        for member_name in sorted(all_members):
            member1_state = members1.get(member_name, {}).get('status.availability-state', 'N/A')
            member2_state = members2.get(member_name, {}).get('status.availability-state', 'N/A')

            match = 'Yes' if member1_state == member2_state else 'No'
            member_full_name = "{} -> {}".format(pool_name, member_name)

            results.append({
                'Type': 'Member',
                'Name': member_full_name,
                col1_header: member1_state,
                col2_header: member2_state,
                'Match': match
            })

            if match == 'No':
                mismatches.append({
                    'Type': 'Member',
                    'Name': member_full_name,
                    col1_header: member1_state,
                    col2_header: member2_state
                })

    # Create DataFrame and save to Excel
    df = pd.DataFrame(results)
    df.to_excel(output_path, index=False)

    # Apply colour formatting to mismatched rows
    wb = load_workbook(output_path)
    ws = wb.active

    # Yellow fill for mismatched rows
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Find the Match column (last column)
    match_col = len(df.columns)

    # Apply formatting (starting from row 2 to skip header)
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=match_col).value == 'No':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.save(output_path)

    # Print results
    print("\nComparison saved to {}".format(output_path))
    print("Compared {} ({}) vs {} ({})".format(host1, time1, host2, time2))

    # Print mismatches to console
    if mismatches:
        print("\n{}".format('=' * 80))
        print("MISMATCHES FOUND: {}".format(len(mismatches)))
        print("{}".format('=' * 80))
        for mismatch in mismatches:
            print("\nType: {}".format(mismatch['Type']))
            print("Name: {}".format(mismatch['Name']))
            print("  {}: {}".format(col1_header, mismatch[col1_header]))
            print("  {}: {}".format(col2_header, mismatch[col2_header]))
    else:
        print("\n{}".format('=' * 80))
        print("NO MISMATCHES FOUND - All states match!")
        print("{}".format('=' * 80))


if __name__ == "__main__":
    pool_type = select_pool_type()

    if pool_type:
        file1, file2 = select_files(pool_type)

        if file1 and file2:
            output_filename = '{}_pool_comparison.xlsx'.format(pool_type)
            compare_pool_states(file1, file2, output_filename)